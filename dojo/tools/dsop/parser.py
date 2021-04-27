import re
from openpyxl import load_workbook

from dojo.models import Finding


class DsopParser:

    def get_scan_types(self):
        return ["DSOP Scan"]

    def get_label_for_scan_types(self, scan_type):
        return scan_type  # no custom label for now

    def get_description_for_scan_types(self, scan_type):
        return "Import XLSX findings from DSOP vulnerability scan pipelines."

    def get_findings(self, file, test):
        book = load_workbook(file)
        items = list()
        self.__parse_disa(test, items, book['OpenSCAP - DISA Compliance'])
        self.__parse_oval(test, items, book['OpenSCAP - OVAL Results'])
        self.__parse_twistlock(test, items, book['Twistlock Vulnerability Results'])
        self.__parse_anchore(test, items, book['Anchore CVE Results'])
        self.__parse_anchore_compliance(test, items, book['Anchore Compliance Results'])
        return items

    def __parse_disa(self, test, items, sheet):
        headers = dict()
        first = True
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                # store the headers
                for i in range(len(row)):
                    headers[row[i]] = i
            else:
                if row[headers['result']] not in ('fail', 'notchecked'):
                    continue
                title = row[headers['title']]
                unique_id = row[headers['ruleid']]
                if row[headers['severity']] == 'unknown':
                    severity = 'Info'
                else:
                    severity = row[headers['severity']].title()
                cve = row[headers['identifiers']]
                references = row[headers['refs']]
                description = row[headers['desc']]
                impact = row[headers['rationale']]
                date = row[headers['scanned_date']]
                tags = "disa"

                finding = Finding(title=title, date=date, cve=cve, severity=severity, description=description,
                            impact=impact, references=references, test=test, unique_id_from_tool=unique_id,
                            static_finding=True, dynamic_finding=False)
                finding.unsaved_tags = tags
                items.append(finding)

    def __parse_oval(self, test, items, sheet):
        severity_pattern = re.compile(r'\((.*)\)')
        headers = dict()
        first = True
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                # store the headers
                for i in range(len(row)):
                    headers[row[i]] = i
            else:
                if not row[headers['result']] or row[headers['result']] in ('false'):
                    continue
                title = row[headers['title']]
                match = severity_pattern.search(title)
                if match:
                    sev = match.group(1)
                    if sev == 'Important':
                        severity = 'High'
                    elif sev == 'Moderate':
                        severity = 'Medium'
                    elif sev == 'None':
                        severity = 'Info'
                    else:
                        severity = sev
                else:
                    severity = 'Info'
                unique_id = row[headers['id']]
                cve = row[headers['ref']]
                tags = "oval"

                finding = Finding(title=title, cve=cve, severity=severity, unique_id_from_tool=unique_id,
                        test=test, static_finding=True, dynamic_finding=False)
                finding.unsaved_tags = tags
                items.append(finding)

    def __parse_twistlock(self, test, items, sheet):
        headers = dict()
        first = True
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                # store the headers
                for i in range(len(row)):
                    headers[row[i]] = i
            else:
                if row[headers['severity']] is None:
                    continue
                cve = row[headers['cve']]
                description = row[headers['desc']]
                mitigation = row[headers['status']]
                url = row[headers['link']]

                component_name = row[headers['packageName']]
                component_version = row[headers['packageVersion']]
                title = '{}: {} - {}'.format(cve, component_name, component_version)
                if row[headers['severity']] == 'important':
                    severity = 'High'
                elif row[headers['severity']] == 'moderate':
                    severity = 'Medium'
                else:
                    severity = row[headers['severity']].title()
                severity_justification = row[headers['vecStr']]
                tags = "twistlock"

                finding = Finding(title=title, cve=cve, url=url, severity=severity, description=description,
                                        component_name=component_name, component_version=component_version,
                                        severity_justification=severity_justification, test=test,
                                        static_finding=True, dynamic_finding=False)
                finding.unsaved_tags = tags
                items.append(finding)

    def __parse_anchore(self, test, items, sheet):
        headers = dict()
        first = True
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                # store the headers
                for i in range(len(row)):
                    headers[row[i]] = i
            else:
                if row[0] is None:
                    continue
                cve = row[headers['cve']]
                severity = row[headers['severity']]
                component = row[headers['package']]
                file_path = row[headers['package_path']]
                mitigation = row[headers['fix']]
                description = "Image affected: {}".format(row[headers['tag']])
                title = '{}: {}'.format(cve, component)
                tags = "anchore"

                finding = Finding(title=title, cve=cve, severity=severity,
                                        mitigation=mitigation, component_name=component,
                                        description=description, test=test,
                                        static_finding=True, dynamic_finding=False,
                                        file_path=file_path)
                finding.unsaved_tags = tags
                items.append(finding)

    def __parse_anchore_compliance(self, test, items, sheet):
        headers = dict()
        first = True
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                # store the headers
                for i in range(len(row)):
                    headers[row[i]] = i
            else:
                if row[headers['policy_id']] != "DoDFileChecks":
                    continue

                if row[headers['gate_action']] == "warn":
                    severity = "Medium"
                elif row[headers['gate_action']] == "stop":
                    severity = "Critical"
                else:
                    severity = "Info"
                severity = severity
                mitigation = "To be investigated"
                description = "Gate: {} (Trigger: {}): {}".format(row[headers['gate']], row[headers['trigger']], row[headers['check_output']])
                title = '{}: {}'.format(row[headers['policy_id']], row[headers['trigger_id']])
                tags = "anchore_compliance"

                finding = Finding(title=title, severity=severity,
                                        mitigation=mitigation,
                                        description=description, test=test,
                                        static_finding=True, dynamic_finding=False)
                finding.unsaved_tags = tags
                items.append(finding)
