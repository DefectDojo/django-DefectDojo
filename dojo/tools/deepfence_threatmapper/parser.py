from openpyxl import load_workbook

from dojo.tools.deepfence_threatmapper.compliance import DeepfenceThreatmapperCompliance
from dojo.tools.deepfence_threatmapper.malware import DeepfenceThreatmapperMalware
from dojo.tools.deepfence_threatmapper.secret import DeepfenceThreatmapperSecret
from dojo.tools.deepfence_threatmapper.vulnerability import DeepfenceThreatmapperVulnerability


class DeepfenceThreatmapperParser:
    def get_scan_types(self):
        return ["Deepfence Threatmapper Report"]

    def get_label_for_scan_types(self, scan_type):
        return scan_type

    def get_description_for_scan_types(self, scan_type):
        return "Deepfence Threatmapper report in XLSX format."

    def get_findings(self, filename, test):
        workbook = load_workbook(filename)
        worksheet = workbook.active
        findings = []
        headers = {}
        first = True
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if first:
                first = False
                for i in range(len(row)):
                    headers[row[i]] = i
            elif (
                ("Rule Name" in headers and "Class" in headers) or
                ("Rule Name" in headers and "Node Type" in headers)
            ):
                findings.append(DeepfenceThreatmapperMalware().get_findings(row, headers, test))
            elif headers.get("Filename") is not None and headers.get("Content") is not None:
                value = DeepfenceThreatmapperSecret().get_findings(row, headers, test)
                if value is not None:
                    findings.append(value)
            elif (
                ("cve_id" in headers and "cve_attack_vector" in headers) or
                ("CVE ID" in headers and "Attack Vector" in headers)
            ):
                findings.append(DeepfenceThreatmapperVulnerability().get_findings(row, headers, test))
            elif (
                ("compliance_check_type" in headers and "test_number" in headers) or
                ("Compliance Standard" in headers and "Control ID" in headers)
            ):
                findings.append(DeepfenceThreatmapperCompliance().get_findings(row, headers, test))
        return findings
