import csv
import logging
import mimetypes
import operator
import re
from datetime import datetime
from functools import reduce
from tempfile import NamedTemporaryFile
from time import strftime

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.utils import NestedObjects
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import DEFAULT_DB_ALIAS
from django.db.models import Count, Q
from django.db.models.query import Prefetch, QuerySet
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, QueryDict, StreamingHttpResponse
from django.shortcuts import get_object_or_404, render
from django.urls import Resolver404, reverse
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views import View
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from openpyxl import Workbook
from openpyxl.styles import Font

import dojo.jira_link.helper as jira_helper
import dojo.risk_acceptance.helper as ra_helper
from dojo.authorization.authorization import user_has_permission_or_403
from dojo.authorization.authorization_decorators import user_is_authorized
from dojo.authorization.roles_permissions import Permissions
from dojo.endpoint.utils import save_endpoints_to_add
from dojo.engagement.queries import get_authorized_engagements
from dojo.engagement.services import close_engagement, reopen_engagement
from dojo.filters import (
    EngagementDirectFilter,
    EngagementDirectFilterWithoutObjectLookups,
    EngagementFilter,
    EngagementFilterWithoutObjectLookups,
    EngagementTestFilter,
    EngagementTestFilterWithoutObjectLookups,
    ProductEngagementsFilter,
    ProductEngagementsFilterWithoutObjectLookups,
)
from dojo.finding.helper import NOT_ACCEPTED_FINDINGS_QUERY
from dojo.finding.views import find_available_notetypes
from dojo.forms import (
    AddFindingsRiskAcceptanceForm,
    CheckForm,
    CredMappingForm,
    DeleteEngagementForm,
    DoneForm,
    EditRiskAcceptanceForm,
    EngForm,
    ImportScanForm,
    JIRAEngagementForm,
    JIRAImportScanForm,
    JIRAProjectForm,
    NoteForm,
    ReplaceRiskAcceptanceProofForm,
    RiskAcceptanceForm,
    TestForm,
    TypedNoteForm,
    UploadThreatForm,
)
from dojo.importers.base_importer import BaseImporter
from dojo.importers.default_importer import DefaultImporter
from dojo.models import (
    Check_List,
    Cred_Mapping,
    Development_Environment,
    Dojo_User,
    Endpoint,
    Engagement,
    Finding,
    Note_Type,
    Notes,
    Product,
    Product_API_Scan_Configuration,
    Risk_Acceptance,
    System_Settings,
    Test,
    Test_Import,
)
from dojo.notifications.helper import create_notification
from dojo.product.queries import get_authorized_products
from dojo.risk_acceptance.helper import prefetch_for_expiration
from dojo.tools.factory import get_scan_types_sorted
from dojo.user.queries import get_authorized_users
from dojo.utils import (
    FileIterWrapper,
    Product_Tab,
    add_breadcrumb,
    add_error_message_to_response,
    add_success_message_to_response,
    async_delete,
    calculate_grade,
    generate_file_response_from_file_path,
    get_cal_event,
    get_page_items,
    get_return_url,
    get_setting,
    get_system_setting,
    handle_uploaded_threat,
    redirect_to_return_url_or_else,
)

logger = logging.getLogger(__name__)


@cache_page(60 * 5)  # cache for 5 minutes
@vary_on_cookie
def engagement_calendar(request):

    if not get_system_setting("enable_calendar"):
        raise Resolver404

    if "lead" not in request.GET or "0" in request.GET.getlist("lead"):
        engagements = get_authorized_engagements(Permissions.Engagement_View)
    else:
        filters = []
        leads = request.GET.getlist("lead", "")
        if "-1" in request.GET.getlist("lead"):
            leads.remove("-1")
            filters.append(Q(lead__isnull=True))
        filters.append(Q(lead__in=leads))
        engagements = get_authorized_engagements(Permissions.Engagement_View).filter(reduce(operator.or_, filters))

    engagements = engagements.select_related("lead")
    engagements = engagements.prefetch_related("product")

    add_breadcrumb(
        title="Engagement Calendar", top_level=True, request=request)
    return render(
        request, "dojo/calendar.html", {
            "caltype": "engagements",
            "leads": request.GET.getlist("lead", ""),
            "engagements": engagements,
            "users": get_authorized_users(Permissions.Engagement_View),
        })


def get_filtered_engagements(request, view):

    if view not in ["all", "active"]:
        msg = f"View {view} is not allowed"
        raise ValidationError(msg)

    engagements = get_authorized_engagements(Permissions.Engagement_View).order_by("-target_start")

    if view == "active":
        engagements = engagements.filter(active=True)

    engagements = engagements.select_related("product", "product__prod_type") \
        .prefetch_related("lead", "tags", "product__tags")

    if System_Settings.objects.get().enable_jira:
        engagements = engagements.prefetch_related(
            "jira_project__jira_instance",
            "product__jira_project_set__jira_instance",
        )

    filter_string_matching = get_system_setting("filter_string_matching", False)
    filter_class = EngagementDirectFilterWithoutObjectLookups if filter_string_matching else EngagementDirectFilter
    return filter_class(request.GET, queryset=engagements)


def get_test_counts(engagements):
    # Get the test counts per engagement. As a separate query, this is much
    # faster than annotating the above `engagements` query.
    return {
        test["engagement"]: test["test_count"]
        for test in Test.objects.filter(
            engagement__in=engagements,
        ).values(
            "engagement",
        ).annotate(
            test_count=Count("engagement"),
        )
    }


def engagements(request, view):

    if not view:
        view = "active"

    filtered_engagements = get_filtered_engagements(request, view)

    engs = get_page_items(request, filtered_engagements.qs, 25)
    product_name_words = sorted(get_authorized_products(Permissions.Product_View).values_list("name", flat=True))
    engagement_name_words = sorted(get_authorized_engagements(Permissions.Engagement_View).values_list("name", flat=True).distinct())

    add_breadcrumb(
        title=f"{view.capitalize()} Engagements",
        top_level=not len(request.GET),
        request=request)

    return render(
        request, "dojo/engagement.html", {
            "engagements": engs,
            "engagement_test_counts": get_test_counts(filtered_engagements.qs),
            "filter_form": filtered_engagements.form,
            "product_name_words": product_name_words,
            "engagement_name_words": engagement_name_words,
            "view": view.capitalize(),
        })


def engagements_all(request):

    products_with_engagements = get_authorized_products(Permissions.Engagement_View)
    products_with_engagements = products_with_engagements.filter(~Q(engagement=None)).distinct()

    # count using prefetch instead of just using 'engagement__set_test_test` to avoid loading all test in memory just to count them
    filter_string_matching = get_system_setting("filter_string_matching", False)
    products_filter_class = ProductEngagementsFilterWithoutObjectLookups if filter_string_matching else ProductEngagementsFilter
    engagement_query = Engagement.objects.annotate(test_count=Count("test__id"))
    filter_qs = products_with_engagements.prefetch_related(
        Prefetch("engagement_set", queryset=products_filter_class(request.GET, engagement_query).qs),
    )

    filter_qs = filter_qs.prefetch_related(
        "engagement_set__tags",
        "prod_type",
        "engagement_set__lead",
        "tags",
    )
    if System_Settings.objects.get().enable_jira:
        filter_qs = filter_qs.prefetch_related(
            "engagement_set__jira_project__jira_instance",
            "jira_project_set__jira_instance",
        )
    filter_class = EngagementFilterWithoutObjectLookups if filter_string_matching else EngagementFilter
    filtered = filter_class(
        request.GET,
        queryset=filter_qs,
    )

    prods = get_page_items(request, filtered.qs, 25)
    prods.paginator.count = sum(len(prod.engagement_set.all()) for prod in prods)
    name_words = products_with_engagements.values_list("name", flat=True)
    eng_words = get_authorized_engagements(Permissions.Engagement_View).values_list("name", flat=True).distinct()

    add_breadcrumb(
        title="All Engagements",
        top_level=not len(request.GET),
        request=request)

    return render(
        request, "dojo/engagements_all.html", {
            "products": prods,
            "filter_form": filtered.form,
            "name_words": sorted(set(name_words)),
            "eng_words": sorted(set(eng_words)),
            "enable_table_filtering": get_system_setting("enable_ui_table_based_searching"),
        })


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def edit_engagement(request, eid):
    engagement = Engagement.objects.get(pk=eid)
    is_ci_cd = engagement.engagement_type == "CI/CD"
    jira_project_form = None
    jira_epic_form = None
    jira_project = None

    if request.method == "POST":
        form = EngForm(request.POST, instance=engagement, cicd=is_ci_cd, product=engagement.product, user=request.user)
        jira_project = jira_helper.get_jira_project(engagement, use_inheritance=False)

        if form.is_valid():
            # first save engagement details
            new_status = form.cleaned_data.get("status")
            engagement.product = form.cleaned_data.get("product")
            engagement = form.save(commit=False)
            if (new_status == "Cancelled" or new_status == "Completed"):
                engagement.active = False
            else:
                engagement.active = True
            engagement.save()
            form.save_m2m()

            messages.add_message(
                request,
                messages.SUCCESS,
                "Engagement updated successfully.",
                extra_tags="alert-success")

            success, jira_project_form = jira_helper.process_jira_project_form(request, instance=jira_project, target="engagement", engagement=engagement, product=engagement.product)
            error = not success

            success, jira_epic_form = jira_helper.process_jira_epic_form(request, engagement=engagement)
            error = error or not success

            if not error:
                if "_Add Tests" in request.POST:
                    return HttpResponseRedirect(
                        reverse("add_tests", args=(engagement.id, )))
                return HttpResponseRedirect(
                    reverse("view_engagement", args=(engagement.id, )))
        else:
            logger.debug(form.errors)

    else:
        form = EngForm(initial={"product": engagement.product}, instance=engagement, cicd=is_ci_cd, product=engagement.product, user=request.user)

        jira_epic_form = None
        if get_system_setting("enable_jira"):
            jira_project = jira_helper.get_jira_project(engagement, use_inheritance=False)
            jira_project_form = JIRAProjectForm(instance=jira_project, target="engagement", product=engagement.product)
            logger.debug("showing jira-epic-form")
            jira_epic_form = JIRAEngagementForm(instance=engagement)

    title = "Edit CI/CD Engagement" if is_ci_cd else "Edit Interactive Engagement"

    product_tab = Product_Tab(engagement.product, title=title, tab="engagements")
    product_tab.setEngagement(engagement)
    return render(request, "dojo/new_eng.html", {
        "product_tab": product_tab,
        "title": title,
        "form": form,
        "edit": True,
        "jira_epic_form": jira_epic_form,
        "jira_project_form": jira_project_form,
        "engagement": engagement,
    })


@user_is_authorized(Engagement, Permissions.Engagement_Delete, "eid")
def delete_engagement(request, eid):
    engagement = get_object_or_404(Engagement, pk=eid)
    product = engagement.product
    form = DeleteEngagementForm(instance=engagement)

    if request.method == "POST":
        if "id" in request.POST and str(engagement.id) == request.POST["id"]:
            form = DeleteEngagementForm(request.POST, instance=engagement)
            if form.is_valid():
                product = engagement.product
                if get_setting("ASYNC_OBJECT_DELETE"):
                    async_del = async_delete()
                    async_del.delete(engagement)
                    message = "Engagement and relationships will be removed in the background."
                else:
                    message = "Engagement and relationships removed."
                    engagement.delete()
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    message,
                    extra_tags="alert-success")
                return HttpResponseRedirect(reverse("view_engagements", args=(product.id, )))

    rels = ["Previewing the relationships has been disabled.", ""]
    display_preview = get_setting("DELETE_PREVIEW")
    if display_preview:
        collector = NestedObjects(using=DEFAULT_DB_ALIAS)
        collector.collect([engagement])
        rels = collector.nested()

    product_tab = Product_Tab(product, title="Delete Engagement", tab="engagements")
    product_tab.setEngagement(engagement)
    return render(request, "dojo/delete_engagement.html", {
        "product_tab": product_tab,
        "engagement": engagement,
        "form": form,
        "rels": rels,
    })


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def copy_engagement(request, eid):
    engagement = get_object_or_404(Engagement, id=eid)
    product = engagement.product
    form = DoneForm()

    if request.method == "POST":
        form = DoneForm(request.POST)
        if form.is_valid():
            engagement_copy = engagement.copy()
            calculate_grade(product)
            messages.add_message(
                request,
                messages.SUCCESS,
                "Engagement Copied successfully.",
                extra_tags="alert-success")
            create_notification(event="engagement_copied",  # TODO: - if 'copy' functionality will be supported by API as well, 'create_notification' needs to be migrated to place where it will be able to cover actions from both interfaces
                                title=_("Copying of %s") % engagement.name,
                                description=f'The engagement "{engagement.name}" was copied by {request.user}',
                                product=product,
                                url=request.build_absolute_uri(reverse("view_engagement", args=(engagement_copy.id, ))),
                                recipients=[engagement.lead],
                                icon="exclamation-triangle")
            return redirect_to_return_url_or_else(request, reverse("view_engagements", args=(product.id, )))
        messages.add_message(
            request,
            messages.ERROR,
            "Unable to copy engagement, please try again.",
            extra_tags="alert-danger")

    product_tab = Product_Tab(product, title="Copy Engagement", tab="engagements")
    return render(request, "dojo/copy_object.html", {
        "source": engagement,
        "source_label": "Engagement",
        "destination_label": "Product",
        "product_tab": product_tab,
        "form": form,
    })


class ViewEngagement(View):

    def get_template(self):
        return "dojo/view_eng.html"

    def get_risks_accepted(self, eng):
        return eng.risk_acceptance.all().select_related("owner").annotate(accepted_findings_count=Count("accepted_findings__id"))

    def get_filtered_tests(
        self,
        request: HttpRequest,
        queryset: list[Test],
        engagement: Engagement,
    ):
        filter_string_matching = get_system_setting("filter_string_matching", False)
        filter_class = EngagementTestFilterWithoutObjectLookups if filter_string_matching else EngagementTestFilter
        return filter_class(request.GET, queryset=queryset, engagement=engagement)

    def get(self, request, eid, *args, **kwargs):
        eng = get_object_or_404(Engagement, id=eid)
        # Make sure the user is authorized
        user_has_permission_or_403(request.user, eng, Permissions.Engagement_View)
        tests = eng.test_set.all().order_by("test_type__name", "-updated")
        default_page_num = 10
        tests_filter = self.get_filtered_tests(request, tests, eng)
        paged_tests = get_page_items(request, tests_filter.qs, default_page_num)
        paged_tests.object_list = prefetch_for_view_tests(paged_tests.object_list)
        prod = eng.product
        risks_accepted = self.get_risks_accepted(eng)
        preset_test_type = None
        network = None
        if eng.preset:
            preset_test_type = eng.preset.test_type.all()
            network = eng.preset.network_locations.all()
        system_settings = System_Settings.objects.get()

        jissue = jira_helper.get_jira_issue(eng)
        jira_project = jira_helper.get_jira_project(eng)

        try:
            check = Check_List.objects.get(engagement=eng)
        except:
            check = None
        notes = eng.notes.all()
        note_type_activation = Note_Type.objects.filter(is_active=True).count()
        if note_type_activation:
            available_note_types = find_available_notetypes(notes)
        form = DoneForm()
        files = eng.files.all()
        form = TypedNoteForm(available_note_types=available_note_types) if note_type_activation else NoteForm()

        creds = Cred_Mapping.objects.filter(
            product=eng.product).select_related("cred_id").order_by("cred_id")
        cred_eng = Cred_Mapping.objects.filter(
            engagement=eng.id).select_related("cred_id").order_by("cred_id")

        add_breadcrumb(parent=eng, top_level=False, request=request)

        title = ""
        if eng.engagement_type == "CI/CD":
            title = " CI/CD"
        product_tab = Product_Tab(prod, title="View" + title + " Engagement", tab="engagements")
        product_tab.setEngagement(eng)
        return render(
            request, self.get_template(), {
                "eng": eng,
                "product_tab": product_tab,
                "system_settings": system_settings,
                "tests": paged_tests,
                "filter": tests_filter,
                "check": check,
                "threat": eng.tmodel_path,
                "form": form,
                "notes": notes,
                "files": files,
                "risks_accepted": risks_accepted,
                "jissue": jissue,
                "jira_project": jira_project,
                "creds": creds,
                "cred_eng": cred_eng,
                "network": network,
                "preset_test_type": preset_test_type,
            })

    def post(self, request, eid, *args, **kwargs):
        eng = get_object_or_404(Engagement, id=eid)
        # Make sure the user is authorized
        user_has_permission_or_403(request.user, eng, Permissions.Engagement_View)
        tests = eng.test_set.all().order_by("test_type__name", "-updated")
        default_page_num = 10

        tests_filter = self.get_filtered_tests(request, tests, eng)
        paged_tests = get_page_items(request, tests_filter.qs, default_page_num)
        # prefetch only after creating the filters to avoid https://code.djangoproject.com/ticket/23771 and https://code.djangoproject.com/ticket/25375
        paged_tests.object_list = prefetch_for_view_tests(paged_tests.object_list)

        prod = eng.product
        risks_accepted = self.get_risks_accepted(eng)
        preset_test_type = None
        network = None
        if eng.preset:
            preset_test_type = eng.preset.test_type.all()
            network = eng.preset.network_locations.all()
        system_settings = System_Settings.objects.get()

        jissue = jira_helper.get_jira_issue(eng)
        jira_project = jira_helper.get_jira_project(eng)

        try:
            check = Check_List.objects.get(engagement=eng)
        except:
            check = None
        notes = eng.notes.all()
        note_type_activation = Note_Type.objects.filter(is_active=True).count()
        if note_type_activation:
            available_note_types = find_available_notetypes(notes)
        form = DoneForm()
        files = eng.files.all()
        user_has_permission_or_403(request.user, eng, Permissions.Note_Add)
        eng.progress = "check_list"
        eng.save()

        if note_type_activation:
            form = TypedNoteForm(request.POST, available_note_types=available_note_types)
        else:
            form = NoteForm(request.POST)
        if form.is_valid():
            new_note = form.save(commit=False)
            new_note.author = request.user
            new_note.date = timezone.now()
            new_note.save()
            eng.notes.add(new_note)
            form = TypedNoteForm(available_note_types=available_note_types) if note_type_activation else NoteForm()
            title = f"Engagement: {eng.name} on {eng.product.name}"
            messages.add_message(request,
                                 messages.SUCCESS,
                                 "Note added successfully.",
                                 extra_tags="alert-success")
        creds = Cred_Mapping.objects.filter(
            product=eng.product).select_related("cred_id").order_by("cred_id")
        cred_eng = Cred_Mapping.objects.filter(
            engagement=eng.id).select_related("cred_id").order_by("cred_id")

        add_breadcrumb(parent=eng, top_level=False, request=request)

        title = ""
        if eng.engagement_type == "CI/CD":
            title = " CI/CD"
        product_tab = Product_Tab(prod, title="View" + title + " Engagement", tab="engagements")
        product_tab.setEngagement(eng)
        return render(
            request, self.get_template(), {
                "eng": eng,
                "product_tab": product_tab,
                "system_settings": system_settings,
                "tests": paged_tests,
                "filter": tests_filter,
                "check": check,
                "threat": eng.tmodel_path,
                "form": form,
                "notes": notes,
                "files": files,
                "risks_accepted": risks_accepted,
                "jissue": jissue,
                "jira_project": jira_project,
                "creds": creds,
                "cred_eng": cred_eng,
                "network": network,
                "preset_test_type": preset_test_type,
            })


def prefetch_for_view_tests(tests):
    prefetched = tests
    if isinstance(tests,
                  QuerySet):  # old code can arrive here with prods being a list because the query was already executed

        prefetched = prefetched.select_related("lead")
        prefetched = prefetched.prefetch_related("tags", "test_type", "notes")
        prefetched = prefetched.annotate(count_findings_test_all=Count("finding__id", distinct=True))
        prefetched = prefetched.annotate(count_findings_test_active=Count("finding__id", filter=Q(finding__active=True), distinct=True))
        prefetched = prefetched.annotate(count_findings_test_active_verified=Count("finding__id", filter=Q(finding__active=True) & Q(finding__verified=True), distinct=True))
        prefetched = prefetched.annotate(count_findings_test_mitigated=Count("finding__id", filter=Q(finding__is_mitigated=True), distinct=True))
        prefetched = prefetched.annotate(count_findings_test_dups=Count("finding__id", filter=Q(finding__duplicate=True), distinct=True))
        prefetched = prefetched.annotate(total_reimport_count=Count("test_import__id", filter=Q(test_import__type=Test_Import.REIMPORT_TYPE), distinct=True))

    else:
        logger.warning("unable to prefetch because query was already executed")

    return prefetched


@user_is_authorized(Engagement, Permissions.Test_Add, "eid")
def add_tests(request, eid):
    eng = Engagement.objects.get(id=eid)
    cred_form = CredMappingForm()
    cred_form.fields["cred_user"].queryset = Cred_Mapping.objects.filter(
        engagement=eng).order_by("cred_id")

    if request.method == "POST":
        form = TestForm(request.POST, engagement=eng)
        cred_form = CredMappingForm(request.POST)
        cred_form.fields["cred_user"].queryset = Cred_Mapping.objects.filter(
            engagement=eng).order_by("cred_id")
        if form.is_valid():
            new_test = form.save(commit=False)
            # set default scan_type as it's used in reimport
            new_test.scan_type = new_test.test_type.name
            new_test.engagement = eng
            try:
                new_test.lead = User.objects.get(id=form["lead"].value())
            except:
                new_test.lead = None

            # Set status to in progress if a test is added
            if eng.status != "In Progress" and eng.active is True:
                eng.status = "In Progress"
                eng.save()

            new_test.save()

            # Save the credential to the test
            if cred_form.is_valid():
                if cred_form.cleaned_data["cred_user"]:
                    # Select the credential mapping object from the selected list and only allow if the credential is associated with the product
                    cred_user = Cred_Mapping.objects.filter(
                        pk=cred_form.cleaned_data["cred_user"].id,
                        engagement=eid).first()

                    new_f = cred_form.save(commit=False)
                    new_f.test = new_test
                    new_f.cred_id = cred_user.cred_id
                    new_f.save()

            messages.add_message(
                request,
                messages.SUCCESS,
                "Test added successfully.",
                extra_tags="alert-success")

            create_notification(
                event="test_added",
                title=f"Test created for {new_test.engagement.product}: {new_test.engagement.name}: {new_test}",
                test=new_test,
                engagement=new_test.engagement,
                product=new_test.engagement.product,
                url=reverse("view_test", args=(new_test.id,)),
                url_api=reverse("test-detail", args=(new_test.id,)),
            )

            if "_Add Another Test" in request.POST:
                return HttpResponseRedirect(
                    reverse("add_tests", args=(eng.id, )))
            if "_Add Findings" in request.POST:
                return HttpResponseRedirect(
                    reverse("add_findings", args=(new_test.id, )))
            if "_Finished" in request.POST:
                return HttpResponseRedirect(
                    reverse("view_engagement", args=(eng.id, )))
    else:
        form = TestForm(engagement=eng)
        form.initial["target_start"] = eng.target_start
        form.initial["target_end"] = eng.target_end
        form.initial["lead"] = request.user
    add_breadcrumb(
        parent=eng, title="Add Tests", top_level=False, request=request)
    product_tab = Product_Tab(eng.product, title="Add Tests", tab="engagements")
    product_tab.setEngagement(eng)
    return render(request, "dojo/add_tests.html", {
        "product_tab": product_tab,
        "form": form,
        "cred_form": cred_form,
        "eid": eid,
        "eng": eng,
    })


class ImportScanResultsView(View):
    def get_template(self) -> str:
        """Returns the template that will be presented to the user"""
        return "dojo/import_scan_results.html"

    def get_development_environment(
        self,
        environment_name: str = "Development",
    ) -> Development_Environment | None:
        """
        Get the development environment in two cases:
        - GET: Environment "Development" by default
        - POST: The label supplied by the user, with Development as a backup
        """
        return Development_Environment.objects.filter(name=environment_name).first()

    def get_engagement_or_product(
        self,
        user: Dojo_User,
        engagement_id: int | None = None,
        product_id: int | None = None,
    ) -> tuple[Engagement, Product, Product | Engagement]:
        """Using the path parameters, either fetch the product or engagement"""
        engagement = product = engagement_or_product = None
        # Get the product if supplied
        # Get the engagement if supplied
        if engagement_id is not None:
            engagement = get_object_or_404(Engagement, id=engagement_id)
            engagement_or_product = engagement
        elif product_id is not None:
            product = get_object_or_404(Product, id=product_id)
            engagement_or_product = product
        else:
            msg = "Either Engagement or Product has to be provided"
            raise Exception(msg)
        # Ensure the supplied user has access to import to the engagement or product
        user_has_permission_or_403(user, engagement_or_product, Permissions.Import_Scan_Result)

        return engagement, product, engagement_or_product

    def get_form(
        self,
        request: HttpRequest,
        **kwargs: dict,
    ) -> ImportScanForm:
        """Returns the default import form for importing findings"""
        if request.method == "POST":
            return ImportScanForm(request.POST, request.FILES, **kwargs)
        return ImportScanForm(**kwargs)

    def get_credential_form(
        self,
        request: HttpRequest,
        engagement: Engagement,
    ) -> CredMappingForm:
        """
        Return a new instance of a form managing credentials. If an engagement
        it present at this time any existing credential objects will be attempted
        to be fetched to populate the form
        """
        if request.method == "POST":
            return CredMappingForm(request.POST)
        # If the engagement is not present, return an empty form
        if engagement is None:
            return CredMappingForm()
        # Otherwise get all creds in the associated engagement
        return CredMappingForm(
            initial={
                "cred_user_queryset": Cred_Mapping.objects.filter(
                    engagement=engagement,
                ).order_by("cred_id"),
            },
        )

    def get_jira_form(
        self,
        request: HttpRequest,
        engagement_or_product: Engagement | Product,
    ) -> tuple[JIRAImportScanForm | None, bool]:
        """Returns a JiraImportScanForm if jira is enabled"""
        jira_form = None
        push_all_jira_issues = False
        # Determine if jira issues should be pushed automatically
        push_all_jira_issues = jira_helper.is_push_all_issues(engagement_or_product)
        # Only return the form if the jira is enabled on this engagement or product
        if jira_helper.get_jira_project(engagement_or_product):
            if request.method == "POST":
                jira_form = JIRAImportScanForm(
                    request.POST,
                    push_all=push_all_jira_issues,
                    prefix="jiraform",
                )
            else:
                jira_form = JIRAImportScanForm(
                    push_all=push_all_jira_issues,
                    prefix="jiraform",
                )
        return jira_form, push_all_jira_issues

    def get_product_tab(
        self,
        product: Product,
        engagement: Engagement,
    ) -> tuple[Product_Tab, dict]:
        """
        Determine how the product tab will be rendered, and what tab will be selected
        as currently active
        """
        custom_breadcrumb = None
        if engagement:
            product_tab = Product_Tab(engagement.product, title="Import Scan Results", tab="engagements")
            product_tab.setEngagement(engagement)
        else:
            custom_breadcrumb = {"", ""}
            product_tab = Product_Tab(product, title="Import Scan Results", tab="findings")
        return product_tab, custom_breadcrumb

    def handle_request(
        self,
        request: HttpRequest,
        engagement_id: int | None = None,
        product_id: int | None = None,
    ) -> tuple[HttpRequest, dict]:
        """
        Process the common behaviors between request types, and then return
        the request and context dict back to be rendered
        """
        user = request.user
        # Get the development environment
        environment = self.get_development_environment()
        # Get the product or engagement from the path parameters
        engagement, product, engagement_or_product = self.get_engagement_or_product(
            user,
            engagement_id=engagement_id,
            product_id=product_id,
        )
        # Get the product tab and any additional custom breadcrumbs
        product_tab, custom_breadcrumb = self.get_product_tab(product, engagement)
        # Get the import form with some initial data in place
        form = self.get_form(
            request,
            environment=environment,
            endpoints=Endpoint.objects.filter(product__id=product_tab.product.id),
            api_scan_configuration=Product_API_Scan_Configuration.objects.filter(product__id=product_tab.product.id),
        )
        # Get the credential mapping form
        cred_form = self.get_credential_form(request, engagement)
        # Get the jira form
        jira_form, push_all_jira_issues = self.get_jira_form(request, engagement_or_product)
        # Return the request and the context
        return request, {
            "user": user,
            "lead": user,
            "form": form,
            "environment": environment,
            "product_tab": product_tab,
            "product": product,
            "engagement": engagement,
            "engagement_or_product": engagement_or_product,
            "custom_breadcrumb": custom_breadcrumb,
            "title": "Import Scan Results",
            "cred_form": cred_form,
            "jform": jira_form,
            "scan_types": get_scan_types_sorted(),
            "push_all_jira_issues": push_all_jira_issues,
        }

    def validate_forms(
        self,
        context: dict,
    ) -> bool:
        """
        Validates each of the forms to ensure all errors from the form
        level are bubbled up to the user first before we process too much
        """
        form_validation_list = []
        if context.get("form") is not None:
            form_validation_list.append(context.get("form").is_valid())
        if context.get("jform") is not None:
            form_validation_list.append(context.get("jform").is_valid())
        if context.get("cred_form") is not None:
            form_validation_list.append(context.get("cred_form").is_valid())
        return all(form_validation_list)

    def create_engagement(
        self,
        context: dict,
    ) -> Engagement:
        """
        Create an engagement if the import was triggered from the product level,
        otherwise, return the existing engagement instead
        """
        # Make sure an engagement does not exist already
        engagement = context.get("engagement")
        if engagement is None:
            engagement = Engagement.objects.create(
                name="AdHoc Import - " + strftime("%a, %d %b %Y %X", timezone.now().timetuple()),
                threat_model=False,
                api_test=False,
                pen_test=False,
                check_list=False,
                active=True,
                target_start=timezone.now().date(),
                target_end=timezone.now().date(),
                product=context.get("product"),
                status="In Progress",
                version=context.get("version"),
                branch_tag=context.get("branch_tag"),
                build_id=context.get("build_id"),
                commit_hash=context.get("commit_hash"),
            )
            # Update the engagement in the context
            context["engagement"] = engagement
        # Return the engagement
        return engagement

    def get_importer(
        self,
        context: dict,
    ) -> BaseImporter:
        """Gets the importer to use"""
        return DefaultImporter(**context)

    def import_findings(
        self,
        context: dict,
    ) -> str | None:
        """Attempt to import with all the supplied information"""
        try:
            importer_client = self.get_importer(context)
            context["test"], _, finding_count, closed_finding_count, _, _, _ = importer_client.process_scan(
                context.pop("scan", None),
            )
            # Add a message to the view for the user to see the results
            add_success_message_to_response(importer_client.construct_imported_message(
                finding_count=finding_count,
                closed_finding_count=closed_finding_count,
            ))
        except Exception as e:
            logger.exception("An exception error occurred during the report import")
            return f"An exception error occurred during the report import: {e}"
        return None

    def process_form(
        self,
        request: HttpRequest,
        form: ImportScanForm,
        context: dict,
    ) -> str | None:
        """Process the form and manipulate the input in any way that is appropriate"""
        # Update the running context dict with cleaned form input
        context.update({
            "scan": request.FILES.get("file", None),
            "scan_date": form.cleaned_data.get("scan_date"),
            "minimum_severity": form.cleaned_data.get("minimum_severity"),
            "active": None,
            "verified": None,
            "scan_type": request.POST.get("scan_type"),
            "tags": form.cleaned_data.get("tags"),
            "version": form.cleaned_data.get("version"),
            "branch_tag": form.cleaned_data.get("branch_tag", None),
            "build_id": form.cleaned_data.get("build_id", None),
            "commit_hash": form.cleaned_data.get("commit_hash", None),
            "api_scan_configuration": form.cleaned_data.get("api_scan_configuration", None),
            "service": form.cleaned_data.get("service", None),
            "close_old_findings": form.cleaned_data.get("close_old_findings", None),
            "apply_tags_to_findings": form.cleaned_data.get("apply_tags_to_findings", False),
            "apply_tags_to_endpoints": form.cleaned_data.get("apply_tags_to_endpoints", False),
            "close_old_findings_product_scope": form.cleaned_data.get("close_old_findings_product_scope", None),
            "group_by": form.cleaned_data.get("group_by", None),
            "create_finding_groups_for_all_findings": form.cleaned_data.get("create_finding_groups_for_all_findings"),
            "environment": self.get_development_environment(environment_name=form.cleaned_data.get("environment")),
        })
        # Create the engagement if necessary
        self.create_engagement(context)
        # close_old_findings_product_scope is a modifier of close_old_findings.
        # If it is selected, close_old_findings should also be selected.
        if close_old_findings_product_scope := form.cleaned_data.get("close_old_findings_product_scope", None):
            context["close_old_findings_product_scope"] = close_old_findings_product_scope
            context["close_old_findings"] = True
        # Save newly added endpoints
        added_endpoints = save_endpoints_to_add(form.endpoints_to_add_list, context.get("engagement").product)
        endpoints_from_form = list(form.cleaned_data["endpoints"])
        context["endpoints_to_add"] = endpoints_from_form + added_endpoints
        # Override the form values of active and verified
        if activeChoice := form.cleaned_data.get("active", None):
            if activeChoice == "force_to_true":
                context["active"] = True
            elif activeChoice == "force_to_false":
                context["active"] = False
        if verifiedChoice := form.cleaned_data.get("verified", None):
            if verifiedChoice == "force_to_true":
                context["verified"] = True
            elif verifiedChoice == "force_to_false":
                context["verified"] = False
        return None

    def process_jira_form(
        self,
        request: HttpRequest,
        form: JIRAImportScanForm,
        context: dict,
    ) -> str | None:
        """
        Process the jira form by first making sure one was supplied
        and then setting any values supplied by the user. An error
        may be returned and will be bubbled up in the form of a message
        """
        # Determine if push all issues is enabled
        push_all_jira_issues = context.get("push_all_jira_issues", False)
        context["push_to_jira"] = push_all_jira_issues or (form and form.cleaned_data.get("push_to_jira"))
        return None

    def process_credentials_form(
        self,
        request: HttpRequest,
        form: CredMappingForm,
        context: dict,
    ) -> str | None:
        """Process the credentials form by creating"""
        if cred_user := form.cleaned_data["cred_user"]:
            # Select the credential mapping object from the selected list and only allow if the credential is associated with the product
            cred_user = Cred_Mapping.objects.filter(
                pk=cred_user.id,
                engagement=context.get("engagement"),
            ).first()
            # Create the new credential mapping object
            new_cred_mapping = form.save(commit=False)
            new_cred_mapping.test = context.get("test")
            new_cred_mapping.cred_id = cred_user.cred_id
            new_cred_mapping.save()
            # update the context
            context["cred_user"] = cred_user
        return None

    def success_redirect(
        self,
        context: dict,
    ) -> HttpResponseRedirect:
        """Redirect the user to a place that indicates a successful import"""
        return HttpResponseRedirect(reverse("view_test", args=(context.get("test").id, )))

    def failure_redirect(
        self,
        context: dict,
    ) -> HttpResponseRedirect:
        """Redirect the user to a place that indicates a failed import"""
        return HttpResponseRedirect(reverse(
            "import_scan_results",
            args=(context.get("engagement", context.get("product")).id, ),
        ))

    def get(
        self,
        request: HttpRequest,
        engagement_id: int | None = None,
        product_id: int | None = None,
    ) -> HttpResponse:
        """Process GET requests for the Import View"""
        # process the request and path parameters
        request, context = self.handle_request(
            request,
            engagement_id=engagement_id,
            product_id=product_id,
        )
        # Render the form
        return render(request, self.get_template(), context)

    def post(
        self,
        request: HttpRequest,
        engagement_id: int | None = None,
        product_id: int | None = None,
    ) -> HttpResponse:
        """Process POST requests for the Import View"""
        # process the request and path parameters
        request, context = self.handle_request(
            request,
            engagement_id=engagement_id,
            product_id=product_id,
        )
        # ensure all three forms are valid first before moving forward
        if not self.validate_forms(context):
            return self.failure_redirect(context)
        # Process the jira form if it is present
        if form_error := self.process_jira_form(request, context.get("jform"), context):
            add_error_message_to_response(form_error)
            return self.failure_redirect(context)
        # Process the import form
        if form_error := self.process_form(request, context.get("form"), context):
            add_error_message_to_response(form_error)
            return self.failure_redirect(context)
        # Kick off the import process
        if import_error := self.import_findings(context):
            add_error_message_to_response(import_error)
            return self.failure_redirect(context)
        # Process the credential form
        if form_error := self.process_credentials_form(request, context.get("cred_form"), context):
            add_error_message_to_response(form_error)
            return self.failure_redirect(context)
        # Otherwise return the user back to the engagement (if present) or the product
        return self.success_redirect(context)


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def close_eng(request, eid):
    eng = Engagement.objects.get(id=eid)
    close_engagement(eng)
    messages.add_message(
        request,
        messages.SUCCESS,
        "Engagement closed successfully.",
        extra_tags="alert-success")
    return HttpResponseRedirect(reverse("view_engagements", args=(eng.product.id, )))


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def reopen_eng(request, eid):
    eng = Engagement.objects.get(id=eid)
    reopen_engagement(eng)
    messages.add_message(
        request,
        messages.SUCCESS,
        "Engagement reopened successfully.",
        extra_tags="alert-success")
    return HttpResponseRedirect(reverse("view_engagements", args=(eng.product.id, )))


"""
Greg:
status: in production
method to complete checklists from the engagement view
"""


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def complete_checklist(request, eid):
    eng = get_object_or_404(Engagement, id=eid)
    try:
        checklist = Check_List.objects.get(engagement=eng)
    except:
        checklist = None

    add_breadcrumb(
        parent=eng,
        title="Complete checklist",
        top_level=False,
        request=request)
    if request.method == "POST":
        tests = Test.objects.filter(engagement=eng)
        findings = Finding.objects.filter(test__in=tests).all()
        form = CheckForm(request.POST, instance=checklist, findings=findings)
        if form.is_valid():
            cl = form.save(commit=False)
            try:
                check_l = Check_List.objects.get(engagement=eng)
                cl.id = check_l.id
                cl.save()
                form.save_m2m()
            except:
                cl.engagement = eng
                cl.save()
                form.save_m2m()
            messages.add_message(
                request,
                messages.SUCCESS,
                "Checklist saved.",
                extra_tags="alert-success")
            return HttpResponseRedirect(
                reverse("view_engagement", args=(eid, )))
    else:
        tests = Test.objects.filter(engagement=eng)
        findings = Finding.objects.filter(test__in=tests).all()
        form = CheckForm(instance=checklist, findings=findings)

    product_tab = Product_Tab(eng.product, title="Checklist", tab="engagements")
    product_tab.setEngagement(eng)
    return render(request, "dojo/checklist.html", {
        "form": form,
        "product_tab": product_tab,
        "eid": eng.id,
        "findings": findings,
    })


@user_is_authorized(Engagement, Permissions.Risk_Acceptance, "eid")
def add_risk_acceptance(request, eid, fid=None):
    eng = get_object_or_404(Engagement, id=eid)
    finding = None
    if fid:
        finding = get_object_or_404(Finding, id=fid)

    if not eng.product.enable_full_risk_acceptance:
        raise PermissionDenied

    if request.method == "POST":
        form = RiskAcceptanceForm(request.POST, request.FILES)
        if form.is_valid():
            # first capture notes param as it cannot be saved directly as m2m
            notes = None
            if form.cleaned_data["notes"]:
                notes = Notes(
                    entry=form.cleaned_data["notes"],
                    author=request.user,
                    date=timezone.now())
                notes.save()

            del form.cleaned_data["notes"]

            try:
                # we sometimes see a weird exception here, but are unable to reproduce.
                # we add some logging in case it happens
                risk_acceptance = form.save()
            except Exception:
                logger.debug(vars(request.POST))
                logger.error(vars(form))
                logger.exception("Creation of Risk Acc. is not possible")
                raise

            # attach note to risk acceptance object now in database
            if notes:
                risk_acceptance.notes.add(notes)

            eng.risk_acceptance.add(risk_acceptance)

            findings = form.cleaned_data["accepted_findings"]

            risk_acceptance = ra_helper.add_findings_to_risk_acceptance(request.user, risk_acceptance, findings)

            messages.add_message(
                request,
                messages.SUCCESS,
                "Risk acceptance saved.",
                extra_tags="alert-success")

            return redirect_to_return_url_or_else(request, reverse("view_engagement", args=(eid, )))
    else:
        risk_acceptance_title_suggestion = f"Accept: {finding}"
        form = RiskAcceptanceForm(initial={"owner": request.user, "name": risk_acceptance_title_suggestion})

    finding_choices = Finding.objects.filter(duplicate=False, test__engagement=eng).filter(NOT_ACCEPTED_FINDINGS_QUERY).order_by("title")

    form.fields["accepted_findings"].queryset = finding_choices
    if fid:
        form.fields["accepted_findings"].initial = {fid}
    product_tab = Product_Tab(eng.product, title="Risk Acceptance", tab="engagements")
    product_tab.setEngagement(eng)

    return render(request, "dojo/add_risk_acceptance.html", {
                  "eng": eng,
                  "product_tab": product_tab,
                  "form": form,
                  })


@user_is_authorized(Engagement, Permissions.Engagement_View, "eid")
def view_risk_acceptance(request, eid, raid):
    return view_edit_risk_acceptance(request, eid=eid, raid=raid, edit_mode=False)


@user_is_authorized(Engagement, Permissions.Risk_Acceptance, "eid")
def edit_risk_acceptance(request, eid, raid):
    return view_edit_risk_acceptance(request, eid=eid, raid=raid, edit_mode=True)


# will only be called by view_risk_acceptance and edit_risk_acceptance
def view_edit_risk_acceptance(request, eid, raid, *, edit_mode=False):
    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)
    eng = get_object_or_404(Engagement, pk=eid)

    if edit_mode and not eng.product.enable_full_risk_acceptance:
        raise PermissionDenied

    risk_acceptance_form = None
    errors = False

    if request.method == "POST":
        # deleting before instantiating the form otherwise django messes up and we end up with an empty path value
        if len(request.FILES) > 0:
            logger.debug("new proof uploaded")
            risk_acceptance.path.delete()

        if "decision" in request.POST:
            old_expiration_date = risk_acceptance.expiration_date
            risk_acceptance_form = EditRiskAcceptanceForm(request.POST, request.FILES, instance=risk_acceptance)
            errors = errors or not risk_acceptance_form.is_valid()
            if not errors:
                logger.debug(f"path: {risk_acceptance_form.cleaned_data['path']}")

                risk_acceptance_form.save()

                if risk_acceptance.expiration_date != old_expiration_date:
                    # risk acceptance was changed, check if risk acceptance needs to be reinstated and findings made accepted again
                    ra_helper.reinstate(risk_acceptance, old_expiration_date)

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    "Risk Acceptance saved successfully.",
                    extra_tags="alert-success")

        if "entry" in request.POST:
            note_form = NoteForm(request.POST)
            errors = errors or not note_form.is_valid()
            if not errors:
                new_note = note_form.save(commit=False)
                new_note.author = request.user
                new_note.date = timezone.now()
                new_note.save()
                risk_acceptance.notes.add(new_note)
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    "Note added successfully.",
                    extra_tags="alert-success")

        if "delete_note" in request.POST:
            note = get_object_or_404(Notes, pk=request.POST["delete_note_id"])
            if note.author.username == request.user.username:
                risk_acceptance.notes.remove(note)
                note.delete()
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    "Note deleted successfully.",
                    extra_tags="alert-success")
            else:
                messages.add_message(
                    request,
                    messages.ERROR,
                    "Since you are not the note's author, it was not deleted.",
                    extra_tags="alert-danger")

        if "remove_finding" in request.POST:
            finding = get_object_or_404(
                Finding, pk=request.POST["remove_finding_id"])

            ra_helper.remove_finding_from_risk_acceptance(request.user, risk_acceptance, finding)

            messages.add_message(
                request,
                messages.SUCCESS,
                "Finding removed successfully from risk acceptance.",
                extra_tags="alert-success")

        if "replace_file" in request.POST:
            replace_form = ReplaceRiskAcceptanceProofForm(
                request.POST, request.FILES, instance=risk_acceptance)

            errors = errors or not replace_form.is_valid()
            if not errors:
                replace_form.save()

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    "New Proof uploaded successfully.",
                    extra_tags="alert-success")
            else:
                logger.error(replace_form.errors)

        if "add_findings" in request.POST:
            add_findings_form = AddFindingsRiskAcceptanceForm(
                request.POST, request.FILES, instance=risk_acceptance)
            errors = errors or not add_findings_form.is_valid()
            if not errors:
                findings = add_findings_form.cleaned_data["accepted_findings"]

                ra_helper.add_findings_to_risk_acceptance(request.user, risk_acceptance, findings)

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    f"Finding{'s' if len(findings) > 1 else ''} added successfully.",
                    extra_tags="alert-success")
        if not errors:
            logger.debug("redirecting to return_url")
            return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))
        logger.error("errors found")

    else:
        if edit_mode:
            risk_acceptance_form = EditRiskAcceptanceForm(instance=risk_acceptance)

    note_form = NoteForm()
    replace_form = ReplaceRiskAcceptanceProofForm(instance=risk_acceptance)
    add_findings_form = AddFindingsRiskAcceptanceForm(instance=risk_acceptance)

    accepted_findings = risk_acceptance.accepted_findings.order_by("numerical_severity")
    fpage = get_page_items(request, accepted_findings, 15)

    unaccepted_findings = Finding.objects.filter(test__in=eng.test_set.all(), risk_accepted=False) \
        .exclude(id__in=accepted_findings).order_by("title")
    add_fpage = get_page_items(request, unaccepted_findings, 25, "apage")
    # on this page we need to add unaccepted findings as possible findings to add as accepted

    add_findings_form.fields[
        "accepted_findings"].queryset = add_fpage.object_list

    add_findings_form.fields["accepted_findings"].widget.request = request
    add_findings_form.fields["accepted_findings"].widget.findings = unaccepted_findings
    add_findings_form.fields["accepted_findings"].widget.page_number = add_fpage.number

    product_tab = Product_Tab(eng.product, title="Risk Acceptance", tab="engagements")
    product_tab.setEngagement(eng)
    return render(
        request, "dojo/view_risk_acceptance.html", {
            "risk_acceptance": risk_acceptance,
            "engagement": eng,
            "product_tab": product_tab,
            "accepted_findings": fpage,
            "notes": risk_acceptance.notes.all(),
            "eng": eng,
            "edit_mode": edit_mode,
            "risk_acceptance_form": risk_acceptance_form,
            "note_form": note_form,
            "replace_form": replace_form,
            "add_findings_form": add_findings_form,
            # 'show_add_findings_form': len(unaccepted_findings),
            "request": request,
            "add_findings": add_fpage,
            "return_url": get_return_url(request),
            "enable_table_filtering": get_system_setting("enable_ui_table_based_searching"),
        })


@user_is_authorized(Engagement, Permissions.Risk_Acceptance, "eid")
def expire_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(prefetch_for_expiration(Risk_Acceptance.objects.all()), pk=raid)
    # Validate the engagement ID exists before moving forward
    get_object_or_404(Engagement, pk=eid)

    ra_helper.expire_now(risk_acceptance)

    return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))


@user_is_authorized(Engagement, Permissions.Risk_Acceptance, "eid")
def reinstate_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(prefetch_for_expiration(Risk_Acceptance.objects.all()), pk=raid)
    eng = get_object_or_404(Engagement, pk=eid)

    if not eng.product.enable_full_risk_acceptance:
        raise PermissionDenied

    ra_helper.reinstate(risk_acceptance, risk_acceptance.expiration_date)

    return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))


@user_is_authorized(Engagement, Permissions.Risk_Acceptance, "eid")
def delete_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)
    eng = get_object_or_404(Engagement, pk=eid)

    ra_helper.delete(eng, risk_acceptance)

    messages.add_message(
        request,
        messages.SUCCESS,
        "Risk acceptance deleted successfully.",
        extra_tags="alert-success")
    return HttpResponseRedirect(reverse("view_engagement", args=(eng.id, )))


@user_is_authorized(Engagement, Permissions.Engagement_View, "eid")
def download_risk_acceptance(request, eid, raid):
    mimetypes.init()
    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)
    # Ensure the risk acceptance is under the supplied engagement
    if not Engagement.objects.filter(risk_acceptance=risk_acceptance, id=eid).exists():
        raise PermissionDenied
    response = StreamingHttpResponse(
        FileIterWrapper(
            open(settings.MEDIA_ROOT + "/" + risk_acceptance.path.name, mode="rb")))
    response["Content-Disposition"] = f'attachment; filename="{risk_acceptance.filename()}"'
    mimetype, _encoding = mimetypes.guess_type(risk_acceptance.path.name)
    response["Content-Type"] = mimetype
    return response


"""
Greg
status: in production
Upload a threat model at the engagement level. Threat models are stored
under media folder
"""


@user_is_authorized(Engagement, Permissions.Engagement_Edit, "eid")
def upload_threatmodel(request, eid):
    eng = Engagement.objects.get(id=eid)
    add_breadcrumb(
        parent=eng,
        title="Upload a threat model",
        top_level=False,
        request=request)

    if request.method == "POST":
        form = UploadThreatForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_threat(request.FILES["file"], eng)
            eng.progress = "other"
            eng.threat_model = True
            eng.save()
            messages.add_message(
                request,
                messages.SUCCESS,
                "Threat model saved.",
                extra_tags="alert-success")
            return HttpResponseRedirect(
                reverse("view_engagement", args=(eid, )))
    else:
        form = UploadThreatForm()
    product_tab = Product_Tab(eng.product, title="Upload Threat Model", tab="engagements")
    return render(request, "dojo/up_threat.html", {
        "form": form,
        "product_tab": product_tab,
        "eng": eng,
    })


@user_is_authorized(Engagement, Permissions.Engagement_View, "eid")
def view_threatmodel(request, eid):
    eng = get_object_or_404(Engagement, pk=eid)
    return generate_file_response_from_file_path(eng.tmodel_path)


@user_is_authorized(Engagement, Permissions.Engagement_View, "eid")
def engagement_ics(request, eid):
    eng = get_object_or_404(Engagement, id=eid)
    start_date = datetime.combine(eng.target_start, datetime.min.time())
    end_date = datetime.combine(eng.target_end, datetime.max.time())
    uid = f"dojo_eng_{eng.id}_{eng.product.id}"
    cal = get_cal_event(
        start_date,
        end_date,
        f"Engagement: {eng.name} ({eng.product.name})",
        (
            f"Set aside for engagement {eng.name}, on product {eng.product.name}. "
            f"Additional detail can be found at {request.build_absolute_uri(reverse('view_engagement', args=(eng.id, )))}"
        ),
        uid,
    )
    output = cal.serialize()
    response = HttpResponse(content=output)
    response["Content-Type"] = "text/calendar"
    response["Content-Disposition"] = f"attachment; filename={eng.name}.ics"
    return response


def get_list_index(full_list, index):
    try:
        element = full_list[index]
    except Exception:
        element = None
    return element


def get_engagements(request):
    url = request.META.get("QUERY_STRING")
    if not url:
        msg = "Please use the export button when exporting engagements"
        raise ValidationError(msg)
    url = url.removeprefix("url=")

    path_items = list(filter(None, re.split(r"/|\?", url)))

    if not path_items or path_items[0] != "engagement":
        msg = "URL is not an engagement view"
        raise ValidationError(msg)

    view = query = None
    if get_list_index(path_items, 1) in ["active", "all"]:
        view = get_list_index(path_items, 1)
        query = get_list_index(path_items, 2)
    else:
        view = "active"
        query = get_list_index(path_items, 1)

    request.GET = QueryDict(query)
    engagements = get_filtered_engagements(request, view).qs
    test_counts = get_test_counts(engagements)

    return engagements, test_counts


def get_excludes():
    return ["is_ci_cd", "jira_issue", "jira_project", "objects", "unaccepted_open_findings"]


def get_foreign_keys():
    return ["build_server", "lead", "orchestration_engine", "preset", "product",
        "report_type", "requester", "source_code_management_server"]


def csv_export(request):
    engagements, test_counts = get_engagements(request)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=engagements.csv"

    writer = csv.writer(response)

    first_row = True
    for engagement in engagements:
        if first_row:
            fields = []
            for key in dir(engagement):
                if key not in get_excludes() and not callable(getattr(engagement, key)) and not key.startswith("_"):
                    fields.append(key)
            fields.append("tests")

            writer.writerow(fields)

            first_row = False
        if not first_row:
            fields = []
            for key in dir(engagement):
                if key not in get_excludes() and not callable(getattr(engagement, key)) and not key.startswith("_"):
                    value = engagement.__dict__.get(key)
                    if key in get_foreign_keys() and getattr(engagement, key):
                        value = str(getattr(engagement, key))
                    if value and isinstance(value, str):
                        value = value.replace("\n", " NEWLINE ").replace("\r", "")
                    fields.append(value)
            fields.append(test_counts.get(engagement.id, 0))

            writer.writerow(fields)

    return response


def excel_export(request):
    engagements, test_counts = get_engagements(request)

    workbook = Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active
    worksheet.title = "Engagements"

    font_bold = Font(bold=True)

    row_num = 1
    for engagement in engagements:
        if row_num == 1:
            col_num = 1
            for key in dir(engagement):
                if key not in get_excludes() and not callable(getattr(engagement, key)) and not key.startswith("_"):
                    cell = worksheet.cell(row=row_num, column=col_num, value=key)
                    cell.font = font_bold
                    col_num += 1
            cell = worksheet.cell(row=row_num, column=col_num, value="tests")
            cell.font = font_bold
            row_num = 2
        if row_num > 1:
            col_num = 1
            for key in dir(engagement):
                if key not in get_excludes() and not callable(getattr(engagement, key)) and not key.startswith("_"):
                    value = engagement.__dict__.get(key)
                    if key in get_foreign_keys() and getattr(engagement, key):
                        value = str(getattr(engagement, key))
                    if value and isinstance(value, datetime):
                        value = value.replace(tzinfo=None)
                    worksheet.cell(row=row_num, column=col_num, value=value)
                    col_num += 1
            worksheet.cell(row=row_num, column=col_num, value=test_counts.get(engagement.id, 0))
        row_num += 1

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(
        content=stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=engagements.xlsx"
    return response
