"""
Regression tests for the V3_FEATURE_LOCATIONS ``Endpoint.__init__`` crash class (issue #15123
and siblings).

When ``V3_FEATURE_LOCATIONS`` is enabled the legacy ``Endpoint`` model is deprecated and
``Endpoint.__init__`` raises ``NotImplementedError``. Findings created/migrated from V2 still
carry legacy ``Endpoint``/``Endpoint_Status`` rows (the migration keeps them as backup), so any
code path that hydrates an ``Endpoint`` under V3 -- directly, via a queryset, or via the
``finding.endpoints`` m2m / ``.endpoint`` FK -- produces a 500.

These tests cover the sites that were guarded/repaired for that:

* JIRA + GitHub issue descriptions -- now render ``finding.locations`` under V3 (the legacy
  ``finding.endpoints`` block is gated out), so a push of a finding that still carries legacy
  endpoint rows must not crash and must show location info.
* CSV + Excel finding exports -- render ``finding.locations.all()`` under V3 (the legacy
  ``finding.endpoints.all()`` iteration is gated out), so exporting a finding that still
  carries legacy endpoint rows must not crash.
* API ``report_generate`` (Product/Engagement) -- ``get_endpoint_ids(Endpoint.objects...)``.
* API ``metadata/batch`` -- the ``endpoint`` parent fetch.
* API ``test_imports`` list -- the ``findings_affected__endpoints`` prefetch on the queryset.
* Batch deduplication -- ``get_finding_models_for_deduplication`` prefetches ``locations``
  under V3, so ``post_process_findings_batch`` no longer hydrates legacy endpoint rows.
"""
import logging
from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth.models import User
from django.test import Client
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from parameterized import parameterized
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient

from dojo.finding.deduplication import get_finding_models_for_deduplication
from dojo.finding.helper import post_process_findings_batch
from dojo.github.services import github_body
from dojo.jira.helper import jira_description
from dojo.location.api.endpoint_compat import V3EndpointStatusCompatibleSerializer
from dojo.location.models import LocationFindingReference, LocationProductReference
from dojo.location.status import FindingLocationStatus, ProductLocationStatus
from dojo.models import (
    IMPORT_CREATED_FINDING,
    Endpoint,
    Endpoint_Status,
    Engagement,
    Finding,
    Product,
    Product_Type,
    Test,
    Test_Import,
    Test_Import_Finding_Action,
    Test_Type,
    UserContactInfo,
)
from dojo.url.models import URL

from .dojo_test_case import DojoTestCase, skip_unless_v3

logger = logging.getLogger(__name__)


@skip_unless_v3
class TestEndpointInitV3(DojoTestCase):

    """
    None of these paths may raise ``NotImplementedError`` when ``V3_FEATURE_LOCATIONS`` is
    enabled and the finding still carries legacy ``Endpoint`` rows.
    """

    def setUp(self):
        super().setUp()

        self.admin = User.objects.create(
            username="test_endpoint_init_v3_admin",
            is_staff=True,
            is_superuser=True,
        )
        UserContactInfo.objects.create(user=self.admin, block_execution=True)

        self.ui_client = Client()
        self.ui_client.force_login(self.admin)

        token, _ = Token.objects.get_or_create(user=self.admin)
        self.api_client = APIClient()
        self.api_client.credentials(HTTP_AUTHORIZATION="Token " + token.key)

        self.system_settings(enable_jira=False)
        self.system_settings(enable_github=False)
        self.system_settings(enable_product_grade=False)

        self.test_type = Test_Type.objects.get_or_create(name="Manual Test")[0]

    # ------------------------------------------------------------------
    # fixtures
    # ------------------------------------------------------------------
    def _make_tree(self, suffix):
        """Product_Type -> Product -> Engagement -> Test -> (active) Finding."""
        product_type = Product_Type.objects.create(name=f"Org {suffix}")
        product = Product.objects.create(
            name=f"Product {suffix}", description="regression fixture", prod_type=product_type,
        )
        engagement = Engagement.objects.create(
            name=f"Eng {suffix}", product=product,
            target_start=timezone.now(), target_end=timezone.now(),
        )
        test = Test.objects.create(
            engagement=engagement, test_type=self.test_type,
            target_start=timezone.now(), target_end=timezone.now(),
        )
        finding = Finding.objects.create(
            test=test, title=f"Finding {suffix}", severity="High", cwe=79,
            description="regression fixture", mitigation="n/a", impact="n/a",
            reporter=self.admin, active=True, verified=True,
        )
        return SimpleNamespace(
            product_type=product_type, product=product,
            engagement=engagement, test=test, finding=finding,
        )

    def _add_legacy_endpoint(self, tree, host):
        """Attach a legacy Endpoint + Endpoint_Status to the finding (as a pre-V3 finding has)."""
        with Endpoint.allow_endpoint_init():
            endpoint = Endpoint(product=tree.product, protocol="https", host=host)
            endpoint.save()
            Endpoint_Status(endpoint=endpoint, finding=tree.finding).save()
        return endpoint

    def _add_location(self, tree, host, status=FindingLocationStatus.Active):
        """Attach a V3 Location to the finding via a LocationFindingReference."""
        url = URL(protocol="https", host=host)
        url.clean()
        saved = URL.bulk_get_or_create([url])
        return LocationFindingReference.objects.create(
            location=saved[0].location, finding=tree.finding, status=status,
        )

    def _add_product_location(self, product, host):
        """Attach a V3 URL Location to the product via a LocationProductReference."""
        url = URL(protocol="https", host=host)
        url.clean()
        saved = URL.bulk_get_or_create([url])
        return LocationProductReference.objects.create(
            location=saved[0].location, product=product, status=ProductLocationStatus.Active,
        )

    # ------------------------------------------------------------------
    # JIRA / GitHub descriptions
    # ------------------------------------------------------------------
    def test_jira_description_renders_locations_under_v3(self):
        """jira_description must render locations (not legacy endpoints) and must not crash."""
        tree = self._make_tree("jira")
        self._add_legacy_endpoint(tree, "legacy-jira.example.com")  # crash trigger under old code
        self._add_location(tree, "loc-jira.example.com")

        description = jira_description(tree.finding)

        self.assertIn("loc-jira.example.com", description)
        # The legacy endpoint block must NOT be rendered under V3.
        self.assertNotIn("legacy-jira.example.com", description)

    def test_github_body_renders_locations_under_v3(self):
        """github_body renders the same template and must not crash under V3."""
        tree = self._make_tree("gh")
        self._add_legacy_endpoint(tree, "legacy-gh.example.com")
        self._add_location(tree, "loc-gh.example.com")

        body = github_body(tree.finding)

        self.assertIn("loc-gh.example.com", body)
        self.assertNotIn("legacy-gh.example.com", body)

    # ------------------------------------------------------------------
    # CSV / Excel exports
    # ------------------------------------------------------------------
    def test_csv_export_renders_locations_under_v3(self):
        """CSV export must render locations (not legacy endpoints) under V3 without crashing."""
        tree = self._make_tree("csv")
        self._add_legacy_endpoint(tree, "legacy-csv.example.com")  # must be ignored under V3
        self._add_location(tree, "loc-csv.example.com")

        response = self.ui_client.get(reverse("csv_export") + f"?url=test/{tree.test.id}")

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("loc-csv.example.com", content)
        self.assertNotIn("legacy-csv.example.com", content)

    def test_excel_export_renders_locations_under_v3(self):
        """Excel export must render locations (not legacy endpoints) under V3 without crashing."""
        tree = self._make_tree("xlsx")
        self._add_legacy_endpoint(tree, "legacy-xlsx.example.com")  # must be ignored under V3
        self._add_location(tree, "loc-xlsx.example.com")

        response = self.ui_client.get(reverse("excel_export") + f"?url=test/{tree.test.id}")

        self.assertEqual(response.status_code, 200)
        cells = [
            str(cell.value)
            for row in load_workbook(BytesIO(response.content)).active.iter_rows()
            for cell in row
        ]
        self.assertTrue(any("loc-xlsx.example.com" in c for c in cells), cells)
        self.assertFalse(any("legacy-xlsx.example.com" in c for c in cells), cells)

    # ------------------------------------------------------------------
    # API report_generate (Product / Engagement)
    # ------------------------------------------------------------------
    def test_api_product_report_generate_renders_locations_under_v3(self):
        """POST /products/{id}/generate_report returns URL locations (compat shape) under V3."""
        tree = self._make_tree("prod-rpt")
        self._add_legacy_endpoint(tree, "legacy-prodrpt.example.com")  # must be ignored under V3
        self._add_product_location(tree.product, "loc-prodrpt.example.com")

        response = self.api_client.post(
            reverse("product-generate-report", args=(tree.product.id,)),
            {}, format="json", HTTP_HOST="testserver",
        )

        self.assertEqual(response.status_code, 200)
        hosts = [ep.get("host") for ep in response.json()["endpoints"]]
        self.assertIn("loc-prodrpt.example.com", hosts)
        self.assertNotIn("legacy-prodrpt.example.com", hosts)

    def test_api_engagement_report_generate_renders_locations_under_v3(self):
        """POST /engagements/{id}/generate_report returns URL locations (compat shape) under V3."""
        tree = self._make_tree("eng-rpt")
        self._add_legacy_endpoint(tree, "legacy-engrpt.example.com")  # must be ignored under V3
        self._add_product_location(tree.product, "loc-engrpt.example.com")

        response = self.api_client.post(
            reverse("engagement-generate-report", args=(tree.engagement.id,)),
            {}, format="json", HTTP_HOST="testserver",
        )

        self.assertEqual(response.status_code, 200)
        hosts = [ep.get("host") for ep in response.json()["endpoints"]]
        self.assertIn("loc-engrpt.example.com", hosts)

    # ------------------------------------------------------------------
    # API test_imports list
    # ------------------------------------------------------------------
    def test_api_test_imports_list_with_legacy_endpoints_under_v3(self):
        """
        GET /test_imports/ must not 500 for a test import whose affected findings still carry
        legacy Endpoint rows.

        Regression: TestImportViewSet.get_queryset() prefetched
        ``findings_affected__endpoints`` unconditionally. Under V3 the prefetch hydrates the
        deprecated Endpoint model, so the paginator raised
        ``NotImplementedError: Endpoint model is deprecated when V3_FEATURE_LOCATIONS is enabled``
        for every caller of the endpoint.
        """
        tree = self._make_tree("test-import")
        self._add_legacy_endpoint(tree, "legacy-testimport.example.com")
        self._add_location(tree, "loc-testimport.example.com")

        test_import = Test_Import.objects.create(test=tree.test, type=Test_Import.IMPORT_TYPE)
        Test_Import_Finding_Action.objects.create(
            test_import=test_import, finding=tree.finding, action=IMPORT_CREATED_FINDING,
        )

        response = self.api_client.get(
            reverse("test_imports-list"), {"test": tree.test.id, "o": "-created", "limit": 1},
        )

        self.assertEqual(
            response.status_code, 200,
            msg=f"expected 200 listing test imports, got {response.status_code}: {response.content[:500]}",
        )
        self.assertEqual([row["id"] for row in response.json()["results"]], [test_import.id])

    # ------------------------------------------------------------------
    # API metadata batch
    # ------------------------------------------------------------------
    def test_api_metadata_batch_with_endpoint_under_v3(self):
        """POST /metadata/batch referencing a legacy endpoint must not 500 under V3."""
        tree = self._make_tree("meta")
        endpoint = self._add_legacy_endpoint(tree, "legacy-meta.example.com")

        response = self.api_client.post(
            reverse("metadata-batch"),
            {
                "product": tree.product.id,
                "endpoint": endpoint.id,
                "metadata": [{"name": "k", "value": "v"}],
            },
            format="json",
        )

        self.assertNotEqual(response.status_code, 500)

    # ------------------------------------------------------------------
    # Batch deduplication
    # ------------------------------------------------------------------
    def test_dedupe_batch_loader_prefetches_locations_under_v3(self):
        """The batch dedupe loader must prefetch locations, never the deprecated endpoints m2m."""
        tree = self._make_tree("dedupe")
        self._add_legacy_endpoint(tree, "legacy-dedupe.example.com")  # crash trigger under old code
        self._add_location(tree, "loc-dedupe.example.com")

        findings = get_finding_models_for_deduplication([tree.finding.id])

        self.assertEqual(len(findings), 1)
        prefetched = sorted(findings[0]._prefetched_objects_cache)
        self.assertIn("locations", prefetched, msg=f"expected locations prefetched, got {prefetched}")
        self.assertNotIn(
            "endpoints", prefetched,
            msg=f"the deprecated endpoints m2m must not be prefetched under V3, got {prefetched}",
        )
        self.assertEqual(
            [ref.location.url.host for ref in findings[0].locations.all()],
            ["loc-dedupe.example.com"],
        )

    def test_post_process_findings_batch_with_legacy_endpoints_under_v3(self):
        """
        The reported crash: post_process_findings_batch loads its findings through the
        dedupe loader, whose endpoints prefetch hydrated legacy Endpoint rows and raised
        NotImplementedError for any finding migrated from V2. Reaching the end of the
        call without an exception is the assertion.
        """
        tree = self._make_tree("batch")
        self._add_legacy_endpoint(tree, "legacy-batch.example.com")
        self._add_location(tree, "loc-batch.example.com")

        post_process_findings_batch(
            [tree.finding.id],
            dedupe_option=False,
            rules_option=False,
            product_grading_option=False,
            issue_updater_option=False,
            push_to_jira=False,
        )

        self.assertTrue(Finding.objects.filter(id=tree.finding.id).exists())

    # ------------------------------------------------------------------
    # endpoint_status compatibility serializer
    # ------------------------------------------------------------------
    # Regression: V3EndpointStatusCompatibleSerializer.get_mitigated returned obj.created.date()
    # (a copy of get_date) instead of the Mitigated status, so every row reported a truthy date.
    @parameterized.expand([
        (FindingLocationStatus.Active, False),
        (FindingLocationStatus.Mitigated, True),
        (FindingLocationStatus.FalsePositive, False),
        (FindingLocationStatus.RiskAccepted, False),
        (FindingLocationStatus.OutOfScope, False),
    ])
    def test_endpoint_status_mitigated_is_a_bool_tracking_the_status(self, status, expected):
        """``mitigated`` must be a bool that is True only for a Mitigated location."""
        tree = self._make_tree(f"mit-{status.value}")
        ref = self._add_location(tree, f"loc-{status.value.lower()}.example.com", status=status)

        data = V3EndpointStatusCompatibleSerializer(ref).data

        self.assertIsInstance(
            data["mitigated"], bool,
            msg=f"expected a bool for status={status.value}, got {type(data['mitigated']).__name__} {data['mitigated']!r}",
        )
        self.assertEqual(
            data["mitigated"], expected,
            msg=f"expected mitigated={expected} for status={status.value}, got {data['mitigated']!r}",
        )

    @parameterized.expand([
        (FindingLocationStatus.Active, False),
        (FindingLocationStatus.Mitigated, True),
    ])
    def test_endpoint_status_mitigated_time_and_by_follow_mitigated(self, status, expected):
        """``mitigated_time``/``mitigated_by`` branch on get_mitigated, so they must follow it."""
        tree = self._make_tree(f"mitby-{status.value}")
        ref = self._add_location(tree, f"loc-mitby-{status.value.lower()}.example.com", status=status)
        ref.auditor = self.admin
        ref.audit_time = timezone.now()
        ref.save()

        data = V3EndpointStatusCompatibleSerializer(ref).data

        self.assertEqual(
            data["mitigated_time"] is not None, expected,
            msg=f"expected mitigated_time set={expected} for status={status.value}, got {data['mitigated_time']!r}",
        )
        self.assertEqual(
            data["mitigated_by"], self.admin.id if expected else None,
            msg=f"expected mitigated_by set={expected} for status={status.value}, got {data['mitigated_by']!r}",
        )
